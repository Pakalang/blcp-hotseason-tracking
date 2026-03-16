#!/usr/bin/env python3
"""
BLCP CW Inlet Tracker — Data Updater
=====================================
วิธีใช้:
  1. วาง Excel ไว้ในโฟลเดอร์เดียวกับ script นี้
  2. รัน:  python update_data.py  หรือดับเบิลคลิก
  3. จะได้ไฟล์ seawater_data.csv (พร้อม commit ขึ้น GitHub)

รองรับ Excel ที่มี columns:
  Date/datetime | inlet | outlet | outfall
"""

import os, glob, csv, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("กำลังติดตั้ง openpyxl ...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ── หาไฟล์ Excel ──────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_files = sorted(
    glob.glob(os.path.join(script_dir, '*.xlsx')) +
    glob.glob(os.path.join(script_dir, '*.xls')),
    key=os.path.getmtime, reverse=True
)

if not xlsx_files:
    print("❌  ไม่พบไฟล์ .xlsx ในโฟลเดอร์นี้")
    input("กด Enter เพื่อปิด...")
    sys.exit(1)

# ถ้ามีหลายไฟล์ ให้ใช้ไฟล์ล่าสุด
target = xlsx_files[0]
print(f"📂  พบไฟล์: {os.path.basename(target)}")

# ── อ่าน Excel ────────────────────────────────────────
wb = openpyxl.load_workbook(target)
ws = wb.active

# detect header row
header = [str(c).lower().strip() if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
print(f"📋  Columns: {header}")

# map columns (ยืดหยุ่น — รองรับชื่อหลากหลาย)
def find_col(keywords):
    for kw in keywords:
        for i, h in enumerate(header):
            if kw in h: return i
    return None

col_dt      = find_col(['date','datetime','time','วัน'])
col_inlet   = find_col(['inlet','cw_inlet','cooling water inlet','อินเล็ต'])
col_outlet  = find_col(['outlet','cw_outlet','cooling water outlet','เอาท์เล็ต'])
col_outfall = find_col(['outfall','discharge','ระบาย'])

if col_dt is None or col_inlet is None:
    print("❌  ไม่พบคอลัมน์ datetime หรือ inlet — ตรวจสอบชื่อหัวตาราง")
    input("กด Enter เพื่อปิด...")
    sys.exit(1)

print(f"✅  Mapped: datetime={col_dt}, inlet={col_inlet}, outlet={col_outlet}, outfall={col_outfall}")

# ── รวบรวมและเรียงข้อมูล ──────────────────────────────
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dt_val = row[col_dt]
    if dt_val is None: continue
    if not isinstance(dt_val, datetime):
        try: dt_val = datetime.fromisoformat(str(dt_val))
        except: continue

    def safe(idx):
        if idx is None: return ''
        v = row[idx]
        return round(float(v), 4) if v is not None else ''

    rows.append([
        dt_val.strftime('%Y-%m-%d %H:%M'),
        safe(col_inlet),
        safe(col_outlet),
        safe(col_outfall),
    ])

rows.sort(key=lambda r: r[0])
print(f"📊  พบข้อมูล {len(rows)} rows ({rows[0][0]} → {rows[-1][0]})")

# ── เขียน CSV ────────────────────────────────────────
out_path = os.path.join(script_dir, 'seawater_data.csv')

# ถ้ามี CSV เดิม → merge (เก็บข้อมูลเก่า + เพิ่มใหม่)
existing = {}
if os.path.exists(out_path):
    with open(out_path, newline='') as f:
        reader = csv.reader(f)
        next(reader, None)
        for r in reader:
            if r: existing[r[0]] = r

before = len(existing)
for r in rows:
    existing[r[0]] = r

merged = sorted(existing.values(), key=lambda r: r[0])
after  = len(merged)

with open(out_path, 'w', newline='') as f:
    w = csv.writer(f)
    w.writerow(['datetime','inlet','outlet','outfall'])
    w.writerows(merged)

print(f"\n✅  บันทึกสำเร็จ: {out_path}")
print(f"   ก่อน merge: {before} rows  →  หลัง merge: {after} rows  (+{after-before} ใหม่)")
print(f"\n📌  ขั้นตอนต่อไป:")
print(f"   1. เปิด GitHub Desktop หรือ git")
print(f"   2. commit ไฟล์ seawater_data.csv")
print(f"   3. push → dashboard จะ update อัตโนมัติ!")
input("\nกด Enter เพื่อปิด...")
